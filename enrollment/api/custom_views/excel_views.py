from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from django.http import FileResponse
from ..utils.services import StudentExcelService #, BillingExcelService
from api.models import Student, AcadTermBilling, Grade, Course, Instructor, Enrollment, BillingList, Receipt
from ..utils.validators import FileValidator
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from django.conf import settings
from django.db.models import Sum
from ..serializers import StudentSerializer, EnrollmentSerializer, ReceiptSerializer
from ..utils.filterer import QuerysetFilter
import os
from django.db import transaction, IntegrityError
import pandas as pd

class StudentExcelAPI(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file uploaded."}, status=400)

        if not FileValidator.is_valid_excel(file):
            return Response({"error": "Invalid file format. Please upload an Excel file."}, status=400)

        try:
            service = StudentExcelService()
            data = service.read(file)
            result = service.process(data)
            return Response({"message": result}, status=200)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


# class BillingExcelAPI(APIView):
#     parser_classes = [MultiPartParser]

#     def post(self, request):
#         file = request.FILES.get("file")
#         if not file:
#             return Response({"error": "No file uploaded."}, status=400)

#         if not FileValidator.is_valid_excel(file):
#             return Response({"error": "Invalid file format. Please upload an Excel file."}, status=400)

#         try:
#             service = BillingExcelService()
#             data = service.read(file)
#             result = service.process(data)
#             return Response({"message": result}, status=200)
#         except Exception as e:
#             return Response({"error": str(e)}, status=500)

class GenerateCORAPI(APIView):
    # API endpoint for generating a Certificate of Registration (COR).

    def get(self, request, student_id):
        try:
            # Fetch the student data
            student = Student.objects.get(id=student_id)
            student_data = StudentSerializer(student).data

            # Fetch enrollments
            enrollments = Enrollment.objects.filter(student=student, school_year=student_data['academic_year'])
            enrollments_data = EnrollmentSerializer(enrollments, many=True).data

            # Fetch billing data
            billing_list = BillingList.objects.all()
            joined_data = []
            for billing in billing_list:
                acad_term_billing = AcadTermBilling.objects.filter(
                    billing=billing,
                    year_level=student_data['year_level'],
                    semester=student_data['semester']
                ).first()

                joined_data.append({
                    "billing_list": {
                        "name": billing.name,
                        "category": billing.category,
                        "price": acad_term_billing.price if acad_term_billing else None,
                        "year_level": acad_term_billing.year_level if acad_term_billing else None,
                        "semester": acad_term_billing.semester if acad_term_billing else None
                    }
                })

            # Calculate total billing price
            assessment_billing_list = BillingList.objects.filter(category='ASSESSMENT')
            total_acad_term_billings = AcadTermBilling.objects.filter(
                billing__in=assessment_billing_list,
                year_level=student_data['year_level'],
                semester=student_data['semester']
            ).aggregate(total_price=Sum('price'))['total_price'] or 0

            # Fetch receipts
            receipts = Receipt.objects.filter(student=student, school_year=student_data['academic_year'])
            receipts_data = ReceiptSerializer(receipts, many=True).data

            # Generate the COR file
            file_path = self.generate_registration_form(student_data, enrollments_data, joined_data, total_acad_term_billings, receipts_data)

            # Return the file as a downloadable response
            return FileResponse(
                open(file_path, 'rb'), as_attachment=True, filename=f"registration_form_{student_id}.xlsx"
            )
        except Student.DoesNotExist:
            return Response({"error": "Student not found."}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    @staticmethod
    def generate_registration_form(student_data, enrollments_data, joined_data, total_acad_term_billings, receipts_data):
        # Create a workbook and sheet
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Registration Form"

        # Add logo image
        logo_path = os.path.join(settings.BASE_DIR, 'static/images/Uni_Logo.png')  # UNIVERSITY LOGO
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width, img.height = 150, 75
            sheet.add_image(img, "A1")

        # Add header information
        header_font = Font(size=12, bold=True)
        sheet.merge_cells("A2:H2")
        sheet["A2"] = "Republic of the Philippines"
        sheet["A2"].font = header_font
        
        sheet.merge_cells("A3:H3")
        sheet["A3"] = "CAVITE STATE UNIVERSITY"
        sheet["A3"].font = Font(size=14, bold=True)

        sheet.merge_cells("A4:H4")
        sheet["A4"] = "Bacoor Campus"
        sheet["A4"].font = header_font

        sheet.merge_cells("A6:H6")
        sheet["A6"] = "REGISTRATION FORM"
        sheet["A6"].font = Font(size=16, bold=True)

        # Add student information
        sheet["A9"] = "Student Number:"
        sheet["B9"] = student_data['id']

        sheet["A10"] = "Student Name:"
        sheet["B10"] = f"{student_data['last_name']}, {student_data['first_name']} {student_data.get('middle_name', '')}"

        sheet["A11"] = "Course:"
        sheet["B11"] = student_data['program']
        sheet["C11"] = "Year:"
        sheet["D11"] = student_data['year_level']

        sheet["A12"] = "Address:"
        address = student_data.get('address', {})
        sheet["B12"] = f"{address.get('street', '')}, {address.get('barangay', '')}, {address.get('city', '')}, {address.get('province', '')}"

        # Add course details header
        sheet["A14"] = "Course Code"
        sheet["B14"] = "Course Title"
        sheet["C14"] = "Units"
        sheet["D14"] = "Time"
        sheet["E14"] = "Day"
        sheet["F14"] = "Room"
        for col in ["A14", "B14", "C14", "D14", "E14", "F14"]:
            sheet[col].font = Font(bold=True)

        # Add course details
        row = 15
        for enrollment in enrollments_data:
            sheet[f"A{row}"] = enrollment['course']['code']
            sheet[f"B{row}"] = enrollment['course']['title']
            sheet[f"C{row}"] = enrollment['course']['lab_units'] + enrollment['course']['lec_units']
            sheet[f"D{row}"] = "TBA"
            sheet[f"E{row}"] = "TBA"
            sheet[f"F{row}"] = "TBA"
            row += 1

        # Add billing details header
        row += 2
        sheet[f"A{row}"] = "Lab Fees"
        sheet[f"B{row}"] = "Other Fees"
        sheet[f"C{row}"] = "Assessment"
        sheet[f"D{row}"] = "Payments"
        for col in [f"A{row}", f"B{row}", f"C{row}", f"D{row}"]:
            sheet[col].font = Font(bold=True)

        # Add billing data
        row += 1
        for billing in joined_data:
            sheet[f"A{row}"] = billing['billing_list']['name']
            sheet[f"B{row}"] = billing['billing_list']['price']
            row += 1

        # Add total billing price
        sheet[f"A{row}"] = "Total Billing Price"
        sheet[f"B{row}"] = total_acad_term_billings

        # Add note section
        row += 2
        sheet.merge_cells(f"A{row}:H{row}")
        sheet[f"A{row}"] = "NOTE: Your slots on the above subjects will be confirmed only upon payment."
        sheet[f"A{row}"].font = Font(bold=True, italic=True)

        # Add additional student information section
        row += 2
        sheet[f"A{row}"] = "Old/New Student:"
        sheet[f"B{row}"] = student_data.get('student_status', '')

        row += 1
        sheet[f"A{row}"] = "Registration Status:"
        sheet[f"B{row}"] = student_data.get('registration_status', '')

        row += 1
        sheet[f"A{row}"] = "Date of Birth:"
        sheet[f"B{row}"] = student_data.get('birth_date', '{Month, Day, Year}')

        row += 1
        sheet[f"A{row}"] = "Gender:"
        sheet[f"B{row}"] = student_data.get('gender', '')

        row += 1
        sheet[f"A{row}"] = "Contact Number:"
        sheet[f"B{row}"] = student_data.get('contact_number', '')

        row += 1
        sheet[f"A{row}"] = "Email Address:"
        sheet[f"B{row}"] = student_data.get('email', '')

        # Add signature section
        row += 2
        sheet.merge_cells(f"A{row}:C{row}")
        sheet[f"A{row}"] = "Student's Signature:"
        sheet[f"A{row}"].font = Font(bold=True)

        # Save the file
        file_path = os.path.join(settings.MEDIA_ROOT, f"registration_form_{student_data['id']}.xlsx")
        wb.save(file_path)
        return file_path

# For imports
class ImportExcelView(APIView):
    parser_classes = [MultiPartParser]

    def find_required_headers(self, data, required_columns):
        """Find the required headers in the DataFrame and return a mapping."""
        found_headers = {col: None for col in required_columns}
        
        # Flatten the DataFrame to search for headers
        for col in data.columns:
            for required_col in required_columns:
                if col.strip().lower() == required_col.lower():
                    found_headers[required_col] = col
                    break
        
        # Check for missing columns
        missing_columns = [col for col, mapped in found_headers.items() if mapped is None]
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")
        
        return found_headers

    def process_students(self, data):
        """Process student data from the uploaded Excel file."""
        required_columns = [
            "id", "first_name", "last_name", "email",
            "contact_number", "program", "gender", "year_level", "status"
        ]
        found_headers = self.find_required_headers(data, required_columns)

        created_count, updated_count = 0, 0

        for _, row in data.iterrows():
            student, created = Student.objects.update_or_create(
                id=row[found_headers["id"]],
                defaults={
                    "first_name": row[found_headers["first_name"]],
                    "last_name": row[found_headers["last_name"]],
                    "email": row[found_headers["email"]],
                    "contact_number": row[found_headers["contact_number"]],
                    "program_id": row[found_headers["program"]],
                    "gender": row[found_headers["gender"]],
                    "year_level": row[found_headers["year_level"]],
                    "status": row[found_headers["status"]],
                }
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

        return f"Successfully imported {created_count} new students and updated {updated_count} existing students."

    def process_grades(self, data):
        """Process grade data from the uploaded Excel file."""
        required_columns = ["student_id", "course_code", "grade", "instructor"]
        found_headers = self.find_required_headers(data, required_columns)

        created_count, updated_count = 0, 0

        for _, row in data.iterrows():
            try:
                # Retrieve the student
                student = Student.objects.get(id=row[found_headers["student_id"]])
                
                # Retrieve the course by code and ensure it belongs to the student's program
                course = Course.objects.get(code=row[found_headers["course_code"]], program=student.program)

                # Attempt to find the instructor or set it to None
                instructor_raw = row[found_headers["instructor"]]
                instructor = None
                if pd.notna(instructor_raw) and isinstance(instructor_raw, str):
                    name_parts = instructor_raw.split()
                    if len(name_parts) >= 2:
                        instructor = Instructor.objects.filter(
                            first_name=name_parts[0],
                            last_name=name_parts[-1]
                        ).first()

                # Check for existing grade before creating/updating
                existing_grade = Grade.objects.filter(student=student, course=course).first()
                if existing_grade:
                    # Update the existing grade
                    existing_grade.grade = row[found_headers["grade"]]
                    existing_grade.instructor = instructor
                    existing_grade.verified = True
                    existing_grade.save()
                    updated_count += 1
                else:
                    # Create a new grade
                    Grade.objects.create(
                        student=student,
                        course=course,
                        grade=row[found_headers["grade"]],
                        instructor=instructor,
                        verified=True
                    )
                    created_count += 1

            except Student.DoesNotExist:
                raise ValueError(f"Student with ID {row[found_headers['student_id']]} does not exist.")
            except Course.DoesNotExist:
                raise ValueError(f"Course with code {row[found_headers['course_code']]} does not exist.")
            except IntegrityError as e:
                # Handle the duplicate entry error
                print(f"Duplicate entry for student {student.id} and course {course.code}: {str(e)}")

        return f"Successfully imported {created_count} new grades and updated {updated_count} existing grades."

    def post(self, request, *args, **kwargs):
        """Handle the POST request for importing an Excel file."""
        file = request.FILES.get('file', None)
        data_type = request.data.get('type', None)

        if not file:
            return Response({"error": "No file uploaded."}, status=400)
        if data_type not in ['students', 'grades']:
            return Response({"error": "Invalid or missing data type. Use 'students' or 'grades'."}, status=400)

        try:
            # Read the Excel file without specifying a header row
            data = pd.read_excel(file)  # Read the file without a specific header

            if data_type == 'students':
                message = self.process_students(data)
            elif data_type == 'grades':
                message = self.process_grades(data)

            return Response({"message": message}, status=200)

        except ValueError as e:
            return Response({"error": str(e)}, status=400)
        except Exception as e:
            return Response({"error": f"An unexpected error occurred: {str(e)}"}, status=500)